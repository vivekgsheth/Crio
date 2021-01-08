from openpyxl import load_workbook

wb2 = load_workbook('t1_data.xlsx')
sheet = wb2.active
ws = wb2['Sheet1']
iplist = []
devices = []
ipad_count = 0
new_l=[]

for tup in tuple(ws.rows):
    if(tup[3].value == "Anushka"):
        iplist.append(tup[2].value)
        devices.append(tup[4].value)
print(iplist)
print(devices)
unique_iplist=set(iplist)
print(unique_iplist)
print(len(set(iplist)))

for item in devices:
    if "iPad" in item:
        ipad_count+=1
print(ipad_count)

# for col in ws.iter_cols(min_col=4, max_col=4, max_row=sheet.max_row):
#     for cell in col:
#         if(cell.value == "Anushka"):
#             iplist.add(col)
# print(iplist)
# for col in ws.iter_cols(min_col=4, max_col=5, max_row=sheet.max_row):
#     for cell in col:
#         if(cell.value == "Anushka"):
#             if 'iPad' in cell.value:
#                 ipad_count+=1
# print(ipad_count)
# print(iplist)
# print(len(iplist))
